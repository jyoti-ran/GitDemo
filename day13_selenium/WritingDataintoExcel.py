import openpyxl

file="C:\mobiledata\jyoti_excel.xlsx"


workbook=openpyxl.load_workbook(file)
sheet=workbook.active   # When you have only one sheet we can write like this
#sheet=workbook["Sheet2"]  # When you have multiple sheet we can write like this

sheet.cell(1,1).value=123
sheet.cell(1,2).value="smith"
sheet.cell(1,3).value="engineer"


sheet.cell(2,1).value=567
sheet.cell(2,2).value="john"
sheet.cell(2,3).value="manager"

sheet.cell(3,1).value=568
sheet.cell(3,2).value="david"
sheet.cell(3,3).value="developer"

sheet.cell(4,1).value=789
sheet.cell(4,2).value="loson"
sheet.cell(4,3).value="lowon"

sheet.cell(5,1).value=987
sheet.cell(5,2).value="aswm"
sheet.cell(5,3).value="location"



workbook.save(file)

